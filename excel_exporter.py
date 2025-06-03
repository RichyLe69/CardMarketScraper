"""
Excel export functionality for CardMarket scraping results.
Handles creating Excel files with proper formatting and organization.
"""

import pandas as pd
import os
from datetime import datetime
from typing import Dict, List, Any
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Exports scraped data to Excel files with proper formatting."""

    def __init__(self):
        pass

    def save_to_excel(self, scraped_data: Dict[str, Dict[str, Any]], list_name: str) -> bool:
        filename = self._generate_filename(list_name)

        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                for sheet_name, sheet_data in scraped_data.items():
                    self._create_sheet(writer, sheet_name, sheet_data)

            print(f"✅ Data saved to {filename}")
            print(f"💡 Each sheet has a clickable URL in cell A1, data starts at row 4")
            return True

        except Exception as e:
            print(f"❌ Error saving Excel file: {e}")
            return False

    def _generate_filename(self, list_name: str) -> str:
        """
        Generate filename with date-based folder structure.
        Creates /output/YYYY-MM-DD/ directory structure.

        Args:
            list_name: Base name for the file

        Returns:
            Complete filepath with date-based folder structure
        """
        # Get current date for folder name
        date_str = datetime.now().strftime("%Y-%m-%d")

        # Create the directory path
        output_dir = os.path.join("output", date_str)

        # Create directories if they don't exist
        os.makedirs(output_dir, exist_ok=True)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y_%m_%d")
        filename = f"{list_name}_{timestamp}.xlsx"

        # Return full path
        return os.path.join(output_dir, filename)
    
    def _create_sheet(self, writer: pd.ExcelWriter, sheet_name: str, sheet_data: Dict[str, Any]):
        """
        Create a single sheet in the Excel file.
        
        Args:
            writer: Excel writer object
            sheet_name: Name of the sheet
            sheet_data: Dictionary containing listings and URL
        """
        listings = sheet_data['listings']
        url = sheet_data['url']
        
        if listings:
            self._create_data_sheet(writer, sheet_name, listings, url)
        else:
            self._create_empty_sheet(writer, sheet_name, url)
    
    def _create_data_sheet(self, writer: pd.ExcelWriter, sheet_name: str, 
                          listings: List[Dict[str, Any]], url: str):
        """
        Create a sheet with listing data.
        
        Args:
            writer: Excel writer object
            sheet_name: Name of the sheet
            listings: List of listing dictionaries
            url: Source URL for the data
        """
        # Create DataFrame
        df = pd.DataFrame(listings)
        
        # Write data starting at row 4 to leave room for URL
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=3)
        
        # Get worksheet and add formatting
        worksheet = writer.sheets[sheet_name]
        self._add_url_to_sheet(worksheet, url)
        self._format_sheet_columns(worksheet)
        
        print(f"  ✅ Created sheet '{sheet_name}' with {len(listings)} listings + URL")
    
    def _create_empty_sheet(self, writer: pd.ExcelWriter, sheet_name: str, url: str):
        """
        Create an empty sheet with headers and URL.
        
        Args:
            writer: Excel writer object
            sheet_name: Name of the sheet
            url: Source URL for the data
        """
        # Create empty DataFrame with headers
        empty_df = pd.DataFrame(columns=[
            'seller_username', 'seller_sales_count', 'condition',
            'condition_badge', 'language', 'edition', 'price', 'quantity'
        ])
        
        # Write headers starting at row 4
        empty_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=3)
        
        # Add URL and formatting
        worksheet = writer.sheets[sheet_name]
        self._add_url_to_sheet(worksheet, url)
        
        print(f"  ⚠️ Created empty sheet '{sheet_name}' with URL")
    
    def _add_url_to_sheet(self, worksheet, url: str):
        """
        Add clickable URL to cell A1.
        
        Args:
            worksheet: Excel worksheet object
            url: URL to add
        """
        worksheet['A1'] = url
        worksheet['A1'].hyperlink = url
        worksheet['A1'].font = Font(color="0563C1", underline="single")
    
    def _format_sheet_columns(self, worksheet):
        """
        Auto-adjust column widths for better readability.
        
        Args:
            worksheet: Excel worksheet object
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set width with padding, but cap at reasonable maximum
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def clean_sheet_name(self, card_name: str) -> str:
        """
        Clean card name to be suitable as Excel sheet name.
        
        Args:
            card_name: Original card name
            
        Returns:
            Cleaned sheet name
        """
        # Replace hyphens with spaces
        sheet_name = card_name.replace('-', ' ')
        
        # Truncate if too long (Excel limit is 31 characters)
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:28] + "..."
        
        return sheet_name
    
    def print_summary(self, scraped_data: Dict[str, Dict[str, Any]]):
        """
        Print summary statistics of the exported data.

        Args:
            scraped_data: Dictionary containing all scraped data
        """
        total_listings = sum(len(sheet_data['listings']) for sheet_data in scraped_data.values())
        total_sheets = len(scraped_data)

        print(f"\n📊 EXPORT SUMMARY:")
        print(f"   Total sheets: {total_sheets}")
        print(f"   Total listings: {total_listings}")

        for sheet_name, sheet_data in scraped_data.items():
            listings = sheet_data['listings']
            if listings:
                with_price = sum(1 for item in listings if item['price'])
                print(f"   {sheet_name}: {len(listings)} listings ({with_price} with price)")
            else:
                print(f"   {sheet_name}: 0 listings")